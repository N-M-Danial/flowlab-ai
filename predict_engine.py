"""
predict_engine.py — Autoregressive prediction + Excel report.

v2 changes (Priority 3 — cold-start fix):
  - predict_date() now accepts base_df and seeds hour-0 lags from
    the previous day's actual hour-23 data from the dataset.
  - If prior-day data is unavailable, falls back gracefully to zeros.
  - Also updated feature vector to include new v2 calendar features.
"""

import os
import warnings
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

from core import (
    TARGET_VOLS, PCU_FACTORS, ALL_FEATS,
    HOURS, MONTHS, WEEKDAYS,
    LOS_BREAKS, LOS_LABELS,
    ZERO_MOTO_ROADS, CLOSED_HOURS_ROAD, CLOSED_HOURS,
    SHEET_ORDER, PRED_DATE,
    compute_los, BASE_DIR,
    get_calendar_features_for_date,
)


# ── Cold-start seed from prior day ────────────────────────────────────────────
def _get_prior_day_seed(road: str, target_date: pd.Timestamp,
                        base_df: pd.DataFrame, cap: float) -> dict:
    """
    Priority 3: Seed hour-0 lags from the actual previous day's hour-22/23 data.
    Returns dict with lag1_vt, lag2_vt, lag3_vt, lag1_vc, lag2_vc, r3_vt, r6_vt, r3_vc.
    Falls back to zeros if data unavailable.
    """
    zero = dict(lag1_vt=0., lag2_vt=0., lag3_vt=0.,
                lag1_vc=0., lag2_vc=0.,
                r3_vt=0., r6_vt=0., r3_vc=0.)

    if base_df is None:
        return zero

    prev_date = target_date - pd.Timedelta(days=1)
    mask = (base_df["road"] == road) & (base_df["date"].dt.normalize() == prev_date.normalize())
    prev = base_df[mask].sort_values("hour")

    if len(prev) < 3:
        return zero

    # Get last few hours of the previous day
    vt_series = (prev[TARGET_VOLS].sum(axis=1)).values   # total volume per hour
    vc_series = prev["vc_ratio"].values

    def _get(arr, idx_from_end):
        idx = len(arr) - 1 - idx_from_end
        return float(arr[idx]) if idx >= 0 else 0.0

    lag1_vt = _get(vt_series, 0)   # hour 23
    lag2_vt = _get(vt_series, 1)   # hour 22
    lag3_vt = _get(vt_series, 2)   # hour 21
    lag1_vc = _get(vc_series, 0)
    lag2_vc = _get(vc_series, 1)

    # Rolling windows: use last 3/6 hours of previous day
    r3_vt = float(np.mean(vt_series[-3:])) if len(vt_series) >= 3 else float(np.mean(vt_series))
    r6_vt = float(np.mean(vt_series[-6:])) if len(vt_series) >= 6 else float(np.mean(vt_series))
    r3_vc = float(np.mean(vc_series[-3:])) if len(vc_series) >= 3 else float(np.mean(vc_series))

    return dict(lag1_vt=lag1_vt, lag2_vt=lag2_vt, lag3_vt=lag3_vt,
                lag1_vc=lag1_vc, lag2_vc=lag2_vc,
                r3_vt=r3_vt, r6_vt=r6_vt, r3_vc=r3_vc)


# ── Autoregressive prediction ─────────────────────────────────────────────────
def predict_date(model, road_params, road_order,
                 target_date=None, base_df=None):
    """
    Predict hourly volumes for every road on target_date.
    base_df: the full historical dataset — used to seed hour-0 lags (Priority 3).
    Returns (predictions dict, summary list).
    """
    target_date = target_date or PRED_DATE
    dow    = target_date.weekday()
    is_we  = int(dow >= 5)
    month  = target_date.month

    # v2 calendar features for this date
    cal = get_calendar_features_for_date(target_date)

    predictions = {}
    summary     = []
    seeded_roads = 0

    for road in road_order:
        p   = road_params[road]
        cap = p["computed_capacity"]

        # ── Priority 3: seed lags from prior day actual data ──────────────────
        seed = _get_prior_day_seed(road, target_date, base_df, cap)
        seeded = any(v != 0.0 for v in seed.values())
        if seeded:
            seeded_roads += 1

        # Initialise history with seeded values so hour-0 has real context
        # We prime with 6 "virtual" prior hours using seed values
        # This gives roll3/roll6 windows real data from the start
        if seeded:
            vt_hist = [seed["lag3_vt"], seed["lag2_vt"], seed["lag1_vt"],
                       seed["lag3_vt"], seed["lag2_vt"], seed["lag1_vt"]]
            vc_hist = [seed["lag2_vc"], seed["lag1_vc"], seed["lag1_vc"],
                       seed["lag2_vc"], seed["lag1_vc"], seed["lag1_vc"]]
        else:
            vt_hist, vc_hist = [], []

        hour_preds = []

        for hour in range(24):
            n = len(vt_hist)

            # Build lag values from history
            lag1_vt = vt_hist[-1] if n >= 1 else 0.0
            lag2_vt = vt_hist[-2] if n >= 2 else 0.0
            lag3_vt = vt_hist[-3] if n >= 3 else 0.0
            lag1_vc = vc_hist[-1] if n >= 1 else 0.0
            lag2_vc = vc_hist[-2] if n >= 2 else 0.0
            r3_vt   = float(np.mean(vt_hist[-3:])) if n >= 1 else 0.0
            r6_vt   = float(np.mean(vt_hist[-6:])) if n >= 1 else 0.0
            r3_vc   = float(np.mean(vc_hist[-3:])) if n >= 1 else 0.0

            # v2 feature vector — 26 features matching ALL_FEATS order:
            # hour_sin, hour_cos, dow_sin, dow_cos,
            # hour, day_of_week, is_weekend, is_peak_morning, is_peak_evening, month,
            # is_public_holiday, is_school_holiday, is_ramadan,
            # road_enc, lanes, design_speed, base_capacity, computed_capacity,
            # vol_total_lag1..3, vc_ratio_lag1..2, roll3/6_vt, roll3_vc
            fv = np.array([[
                cal["hour_sin"][hour], cal["hour_cos"][hour],
                cal["dow_sin"],        cal["dow_cos"],
                hour, dow, is_we,
                int(hour in {7, 8, 9}), int(hour in {17, 18, 19}),
                month,
                cal["is_public_holiday"], cal["is_school_holiday"], cal["is_ramadan"],
                p["road_enc"], p["lanes"], p["design_speed"],
                p["base_capacity"], cap,
                lag1_vt, lag2_vt, lag3_vt,
                lag1_vc, lag2_vc,
                r3_vt, r6_vt, r3_vc,
            ]])

            counts = list(np.clip(model.predict(fv)[0], 0, None))

            # Structural zero clamps
            if road in ZERO_MOTO_ROADS:
                counts[1] = 0.0
            if road == CLOSED_HOURS_ROAD and hour in CLOSED_HOURS:
                counts = [0.0] * 6

            counts = [round(float(c)) for c in counts]
            hour_preds.append(counts)

            total = sum(counts)
            pcu   = sum(v * f for v, f in zip(counts, PCU_FACTORS))
            vc    = round(pcu / cap, 6) if cap > 0 else 0.0
            vt_hist.append(total)
            vc_hist.append(vc)

        predictions[road] = hour_preds

        vcs       = [sum(v * f for v, f in zip(r, PCU_FACTORS)) / cap
                     if cap > 0 else 0 for r in hour_preds]
        peak_vc   = max(vcs)
        peak_hour = HOURS[int(np.argmax(vcs))]
        avg_vc    = float(np.mean(vcs))
        daily     = sum(sum(r) for r in hour_preds)
        summary.append({
            "road":      road,
            "daily":     daily,
            "peak_vc":   round(peak_vc, 4),
            "peak_hour": peak_hour,
            "avg_vc":    round(avg_vc, 4),
            "peak_los":  compute_los(peak_vc),
            "avg_los":   compute_los(avg_vc),
            "seeded":    seeded,
        })

    return predictions, summary


# ── Prediction drift monitoring ────────────────────────────────────────────────
def rolling_vc_error(log_path, window=48):
    if not os.path.exists(log_path):
        return {"status": "no_log", "rolling_mae": None}
    df = pd.read_csv(log_path)
    recent = df.tail(window)
    if len(recent) < 2:
        return {"status": "insufficient_data", "rolling_mae": None}
    mae = float(np.mean(np.abs(recent["pred_vc"] - recent["actual_vc"])))
    return {"status": "alert" if mae > 0.15 else "ok",
            "rolling_mae": round(mae, 5),
            "window": window, "n_samples": len(recent)}


def log_prediction(pred_vc, actual_vc, road, hour, log_path=None):
    log_path = log_path or os.path.join(BASE_DIR, "logs", "prediction_log.csv")
    row = pd.DataFrame([{"timestamp": pd.Timestamp.utcnow().isoformat(),
                          "road": road, "hour": hour,
                          "pred_vc": pred_vc, "actual_vc": actual_vc}])
    row.to_csv(log_path, mode="a", header=not os.path.exists(log_path), index=False)


# ── Excel report (identical output to original) ────────────────────────────────
def _thin_border():
    s = Side(border_style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply(cell, value=None, bold=False, fill=None, align="center",
           valign="center", wrap=False, border=True, fmt=None,
           size=10, color="000000"):
    if value is not None:
        cell.value = value
    cell.font      = Font(bold=bold, size=size, color=color, name="Calibri")
    if fill:
        hx = fill if len(fill) == 8 else "FF" + fill
        cell.fill = PatternFill("solid", fgColor=hx)
    cell.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        cell.border = _thin_border()
    if fmt:
        cell.number_format = fmt


def _build_sheet(ws, road_name, params, volumes, target_date):
    lanes    = params["lanes"]
    speed    = params["design_speed"]
    base_cap = params["base_capacity"]
    adj      = params["adj_factor"]

    for rng in ["A1:L1","A2:L2","A3:C3","E3:F3","H3:L3",
                "A4:B4","A5:B5","A6:B6","A7:B7","A8:B8","A9:B9",
                "J4:L4","J5:L5","J6:L6","J7:L7","J8:L8","J9:L9"]:
        ws.merge_cells(rng)

    ws.row_dimensions[1].height  = 31.5
    ws.row_dimensions[2].height  = 9.0
    for r in range(3, 10):
        ws.row_dimensions[r].height = 18.0
    ws.row_dimensions[10].height = 6.0
    ws.row_dimensions[11].height = 29.25
    for r in range(12, 36):
        ws.row_dimensions[r].height = 16.5
    ws.row_dimensions[36].height = 19.5

    for col, w in {"A": 11, "B": 14, "C": 13.43, "D": 7, "E": 12.43,
                   "F": 11, "G": 7, "H": 9, "I": 10, "J": 13, "K": 8, "L": 7}.items():
        ws.column_dimensions[col].width = w

    title = (f"{road_name}  ·  Hourly Traffic Volume & LOS Report  ·  "
             f"{target_date.day} {MONTHS[target_date.month-1]} {target_date.year} "
             f"({WEEKDAYS[target_date.weekday()]})")
    _apply(ws["A1"], title, bold=True, fill="1F3864", color="FFFFFF", size=12)
    _apply(ws["A3"], "  ROAD PARAMETERS ",       bold=True, fill="2E75B6", color="FFFFFF")
    _apply(ws["E3"], "  PCU EQUIVALENTS ",        bold=True, fill="2E75B6", color="FFFFFF")
    _apply(ws["H3"], "  LEVEL OF SERVICE LEGEND", bold=True, fill="2E75B6", color="FFFFFF")

    param_rows = [
        ("Lane Count (both directions)", lanes),
        ("Design Speed (km/h)",          speed),
        ("Base Capacity (PCU/lane/hr)",  base_cap),
        ("Adjustment Factor",            adj),
        (None, None),
        ("Computed Capacity (PCU/hr)  ▶","=$C$6*$C$4*$C$7"),
    ]
    pcu_rows = [("Car",1.0),("Motorcycles",0.5),("Van",1.5),
                ("Medium Lorry",2.0),("Heavy Lorry",3.0),("Bus",3.0)]
    los_rows = [("A","≤ 0.60","Free Flow"),("B","≤ 0.70","Reasonable Flow"),
                ("C","≤ 0.80","Stable Flow"),("D","≤ 0.90","Near Capacity"),
                ("E","≤ 1.00","At Capacity"),("F","> 1.00","Over Capacity")]

    for i, rn in enumerate(range(4, 10)):
        lbl, val = param_rows[i]
        if lbl:
            _apply(ws.cell(rn, 1), lbl, fill="EBF3FB", align="left")
        ws.cell(rn, 2).border = _thin_border()
        if val is not None:
            _apply(ws.cell(rn, 3), val, bold=True, fill="FFF2CC")
        pl, pv = pcu_rows[i]
        _apply(ws.cell(rn, 5), pl, fill="EBF3FB", align="left")
        _apply(ws.cell(rn, 6), pv, bold=True, fill="FFF2CC")
        lg, lt, ld = los_rows[i]
        _apply(ws.cell(rn, 8), lg, bold=True, fill="EBF3FB")
        _apply(ws.cell(rn, 9), lt, fill="F2F2F2")
        _apply(ws.cell(rn, 10), ld, fill="F2F2F2", align="left")

    for c, h in enumerate(["Time","Car","Motorcycles","Van","Medium\nLorry",
                            "Heavy\nLorry","Bus","Total\nVehicles","Volume\n(PCU)",
                            "Capacity\n(PCU)","V/C\nRatio","LOS\nGrade"], 1):
        _apply(ws.cell(11, c), h, bold=True, fill="D6E4F7", wrap=True)

    for hour, (hour_lbl, counts) in enumerate(zip(HOURS, volumes)):
        r = 12 + hour
        _apply(ws.cell(r, 1), hour_lbl)
        for ci, v in enumerate(counts, 2):
            _apply(ws.cell(r, ci), int(v), fmt="#,##0")
        _apply(ws.cell(r, 8),  f"=B{r}+C{r}+D{r}+E{r}+F{r}+G{r}", fmt="#,##0")
        _apply(ws.cell(r, 9),
               f"=B{r}*$F$4+C{r}*$F$5+D{r}*$F$6+E{r}*$F$7+F{r}*$F$8+G{r}*$F$9",
               fmt="#,##0")
        _apply(ws.cell(r, 10), "=$C$9",                     fill="F5F5F5", fmt="#,##0")
        _apply(ws.cell(r, 11), f"=IF(J{r}>0,I{r}/J{r},0)",  fmt="0.000")
        _apply(ws.cell(r, 12),
               f'=IF(K{r}<=0.6,"A",IF(K{r}<=0.7,"B",IF(K{r}<=0.8,"C",'
               f'IF(K{r}<=0.9,"D",IF(K{r}<=1,"E","F")))))', bold=True)

    _apply(ws.cell(36, 1), "TOTAL", bold=True, fill="BDD7EE")
    for c in range(2, 10):
        _apply(ws.cell(36, c),
               f"=SUM({get_column_letter(c)}12:{get_column_letter(c)}35)",
               bold=True, fill="BDD7EE", fmt="#,##0")
    for c in [10, 11, 12]:
        _apply(ws.cell(36, c), "—", bold=True, fill="BDD7EE")

    ws.freeze_panes = "A12"


def build_excel(predictions, road_params, target_date=None, out_path=None):
    target_date = target_date or PRED_DATE
    date_str    = target_date.strftime("%Y%m%d")
    out_path    = out_path or os.path.join(BASE_DIR, f"Traffic_LOS_{date_str}.xlsx")
    wb = Workbook()
    wb.remove(wb.active)
    for road in SHEET_ORDER:
        if road not in predictions:
            continue
        ws = wb.create_sheet(title=road)
        _build_sheet(ws, road, road_params[road], predictions[road], target_date)
    wb.save(out_path)
    return out_path
